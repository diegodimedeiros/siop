import io

from django.http import HttpResponse
from django.utils import timezone

from core.utils.formatters import to_export_text, user_display
from core.utils.helpers import build_rows


def export_generic_excel(
    request,
    queryset,
    *,
    filename_prefix,
    sheet_title,
    document_title,
    document_subject,
    headers,
    row_getters,
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse(
            "openpyxl não está instalado. Execute: pip install openpyxl",
            status=500,
        )

    now_local = timezone.localtime(timezone.now())
    filename = f"{filename_prefix}_{now_local.strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    creator = user_display(getattr(request, "user", None)) or "Sistema"

    wb.properties.title = document_title
    wb.properties.subject = document_subject
    wb.properties.creator = creator
    wb.properties.category = "Relatório"
    wb.properties.comments = f"Gerado em {now_local.strftime('%d/%m/%Y %H:%M')} pelo SIOP"
    wb.properties.created = now_local
    wb.properties.modified = now_local

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_fill_green = PatternFill(start_color="E1F8EE", end_color="E1F8EE", fill_type="solid")
    body_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, row in enumerate(build_rows(queryset, row_getters), start=2):
        safe_row = [to_export_text(value) for value in row]
        ws.append(safe_row)

        fill = body_fill_green if row_idx % 2 == 0 else body_fill_white
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response